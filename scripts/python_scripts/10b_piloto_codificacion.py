"""Paso 10b: Extracción de mensajes para codificación piloto manual
===================================================================
Genera un Excel con los mensajes de los tres grupo-semana seleccionados
para la codificación piloto de los 7 indicadores de Dedios-Sanguineti et
al. (2025), más una plantilla de codificación y una guía de indicadores.

Grupos piloto seleccionados (criterio: mayor número de marcadores
lingüísticos explícitos de respuesta P-P, diversidad de ciudad/género/tema):

  - GMV9 semana 6  | Valledupar | Mujeres | "Acompañando las emociones"
  - GMB3 semana 4  | Bogotá     | Mujeres | "Crianza integral"
  - GHV9 semana 3  | Valledupar | Hombres | "Conociendo el crecimiento"

La unidad de codificación es la "interacción temática": una cadena de
mensajes sobre el mismo tema dentro de la sesión. El codificador la
delimita manualmente y luego aplica los 7 indicadores (+ 1 específico
de Apapachar) a cada interacción.

Input:
  - data/clean/mensajes_preprocesados.parquet
  - outputs/tables/10a_cadenas_sesion.csv

Output:
  - outputs/tables/10b_piloto_codificacion.xlsx

Para la codificación a escala (Chunks 1–6) ver 10c_codificacion.py.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data" / "clean"
TABLES_DIR = ROOT / "outputs" / "tables"

GRUPOS_PILOTO = [
    (
        "GMV9",
        6,
        "Valledupar",
        "Mujeres",
        "Acompañando las emociones de nuestras niñas y niños",
    ),
    (
        "GMB3",
        4,
        "Bogotá",
        "Mujeres",
        "Crianza integral: estrategias nuevas para fomentar el aprendizaje",
    ),
    (
        "GHV9",
        3,
        "Valledupar",
        "Hombres",
        "Conociendo el crecimiento de los niños: aprender sobre el desarrollo",
    ),
]

# ---------------------------------------------------------------------------
# Guía de indicadores
# ---------------------------------------------------------------------------
GUIA_INDICADORES = [
    {
        "Nivel": "Stance-only",
        "N": "I1",
        "Indicador": "Emergencia de posturas",
        "Definición (Dedios et al.)": (
            "Un participante comparte su punto de vista, opinión o experiencia "
            "sobre el tema sin que genere respuesta de otro participante."
        ),
        "Adaptación Apapachar": (
            "Postura sobre una práctica de crianza, sobre corresponsabilidad "
            "en el hogar, o sobre una experiencia familiar pasada o presente. "
            "Ejemplo: 'En mi casa siempre le pegamos a los hijos cuando no obedecen.'"
        ),
        "Código en plantilla": "I1_stance",
    },
    {
        "Nivel": "Interacción básica",
        "N": "I2",
        "Indicador": "Consenso",
        "Definición (Dedios et al.)": (
            "Dos o más participantes expresan acuerdo explícito con la postura "
            "de otro participante."
        ),
        "Adaptación Apapachar": (
            "Un participante valida o refuerza lo que dijo otro. Marcadores: "
            "'yo también', 'exactamente', 'estoy de acuerdo', 'como dijo [nombre]', "
            "'me pasa lo mismo'."
        ),
        "Código en plantilla": "I2_consenso",
    },
    {
        "Nivel": "Interacción básica",
        "N": "I3",
        "Indicador": "Desacuerdo",
        "Definición (Dedios et al.)": (
            "Un participante desafía o contradice explícitamente la postura "
            "de otro participante."
        ),
        "Adaptación Apapachar": (
            "Un participante matiza o cuestiona lo que dijo otro. Puede ser "
            "suave ('pero en mi caso...', 'no sé si estoy de acuerdo') o más "
            "directo. Incluye tensión entre crianza tradicional y nueva."
        ),
        "Código en plantilla": "I3_desacuerdo",
    },
    {
        "Nivel": "Interacción básica",
        "N": "I4",
        "Indicador": "Cambio de posición",
        "Definición (Dedios et al.)": (
            "Un participante modifica una postura que había expresado previamente "
            "en la misma discusión."
        ),
        "Adaptación Apapachar": (
            "Incluye también cuando un participante reporta que cambió una "
            "práctica en casa como resultado de la discusión en el grupo. "
            "Ejemplo: 'Antes le gritaba pero ahora que hablamos de esto lo "
            "intento diferente.' Este es el equivalente de shifting para "
            "intervenciones de cambio conductual."
        ),
        "Código en plantilla": "I4_cambio_posicion",
    },
    {
        "Nivel": "Interacción compleja",
        "N": "I5",
        "Indicador": "Construcción descriptiva de normalidad",
        "Definición (Dedios et al.)": (
            "Los participantes negocian colectivamente qué comportamientos o "
            "situaciones son 'normales' o 'típicas' en su contexto."
        ),
        "Adaptación Apapachar": (
            "Participantes co-construyen qué es 'normal' en crianza. Ej: "
            "'En Colombia siempre ha sido así', 'la mayoría de padres lo hace', "
            "'en mi familia todos lo hacíamos'. Puede confirmar o cuestionar "
            "la norma cultural."
        ),
        "Código en plantilla": "I5_normalidad",
    },
    {
        "Nivel": "Interacción compleja",
        "N": "I6",
        "Indicador": "Construcción moral",
        "Definición (Dedios et al.)": (
            "Los participantes negocian juicios morales, evaluando "
            "comportamientos como correctos/incorrectos, buenos/malos."
        ),
        "Adaptación Apapachar": (
            "Evaluación moral de prácticas de crianza, corresponsabilidad, "
            "o violencia intrafamiliar. Ejemplo: 'eso está mal', 'no está "
            "bien que los niños reciban golpes', 'el hombre también debe "
            "participar en la crianza'. Puede ser conflictivo o de consenso moral."
        ),
        "Código en plantilla": "I6_moral",
    },
    {
        "Nivel": "Interacción compleja",
        "N": "I7",
        "Indicador": "Identidades compartidas",
        "Definición (Dedios et al.)": (
            "Los participantes co-construyen una identidad colectiva como grupo."
        ),
        "Adaptación Apapachar": (
            "Uso del 'nosotros' inclusivo que define al grupo. Ejemplos: "
            "'nosotras como madres sabemos que...', 'los que estamos aquí "
            "queremos mejorar', 'como papás entendemos...'. Diferente del "
            "'nosotros' familiar ('en mi familia')."
        ),
        "Código en plantilla": "I7_identidad",
    },
    {
        "Nivel": "Específico Apapachar",
        "N": "I8",
        "Indicador": "Adopción de práctica reportada",
        "Definición (Dedios et al.)": "N/A — indicador nuevo propuesto para intervenciones",
        "Adaptación Apapachar": (
            "Un participante reporta explícitamente haber aplicado algo del "
            "programa en casa durante la semana. Evidencia directa de "
            "transferencia. Ejemplo: 'Esta semana hice lo del juego con mi "
            "hija y le gustó mucho', 'probé hablarle con calma y sí funcionó'."
        ),
        "Código en plantilla": "I8_adopcion_practica",
    },
]

# ---------------------------------------------------------------------------
# Columnas de la plantilla de codificación
# ---------------------------------------------------------------------------
COLS_PLANTILLA = [
    "grupo",
    "semana",
    "ciudad",
    "genero",
    "interaccion_id",
    "datetime_inicio",
    "datetime_fin",
    "n_mensajes",
    "ids_participantes",
    "resumen_tematica",
    "I1_stance",
    "I2_consenso",
    "I3_desacuerdo",
    "I4_cambio_posicion",
    "I5_normalidad",
    "I6_moral",
    "I7_identidad",
    "I8_adopcion_practica",
    "nivel_interaccion",
    "citas_relevantes",
    "notas",
]

NIVEL_INSTRUCCION = (
    "nivel_interaccion: escribir 'stance-only' si solo I1=1; "
    "'basica' si I2, I3 o I4 ≥ 1; 'compleja' si I5, I6 o I7 ≥ 1. "
    "Si hay varios niveles, poner el más alto."
)

# ---------------------------------------------------------------------------
# ITs codificadas — Piloto (GMV9_s6, GMB3_s4, GHV9_s3)
# ---------------------------------------------------------------------------
PILOTO_ITS = [
    {
        "grupo": "GMV9",
        "semana": 6,
        "ciudad": "Valledupar",
        "genero": "Mujer",
        "interaccion_id": "IT_01",
        "resumen_tematica": "Reacciones de cuidadoras cuando el niño esta triste o preocupado",
        "I1_stance": 4,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 0,
        "I6_moral": 0,
        "I7_identidad": 1,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GMV9",
        "semana": 6,
        "ciudad": "Valledupar",
        "genero": "Mujer",
        "interaccion_id": "IT_02",
        "resumen_tematica": 'Opinion de cuidadoras sobre el refrán tradicional "Los hombres no llorán"',
        "I1_stance": 1,
        "I2_consenso": 1,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 1,
        "I6_moral": 1,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GMV9",
        "semana": 6,
        "ciudad": "Valledupar",
        "genero": "Mujer",
        "interaccion_id": "IT_03",
        "resumen_tematica": "Opinion de cuidadoras sobre emociones que pueden sentir y expresar o no los hombres",
        "I1_stance": 2,
        "I2_consenso": 0,
        "I3_desacuerdo": 1,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 1,
        "I6_moral": 1,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GMV9",
        "semana": 6,
        "ciudad": "Valledupar",
        "genero": "Mujer",
        "interaccion_id": "IT_04",
        "resumen_tematica": "¿Hay emociones buenas y malas?",
        "I1_stance": 1,
        "I2_consenso": 1,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 1,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GMV9",
        "semana": 6,
        "ciudad": "Valledupar",
        "genero": "Mujer",
        "interaccion_id": "IT_05",
        "resumen_tematica": "Reflexión de cierre sobre lo aprendido en la semana",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 0,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "stance-only",
    },
    {
        "grupo": "GMB3",
        "semana": 4,
        "ciudad": "Bogotá",
        "genero": "Mujer",
        "interaccion_id": "IT_01",
        "resumen_tematica": "Estrategia de redirigir la atención en pataletas",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 0,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "stance-only",
    },
    {
        "grupo": "GMB3",
        "semana": 4,
        "ciudad": "Bogotá",
        "genero": "Mujer",
        "interaccion_id": "IT_02",
        "resumen_tematica": "Estrategia de dar opciones según la época y el niño",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 0,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "stance-only",
    },
    {
        "grupo": "GMB3",
        "semana": 4,
        "ciudad": "Bogotá",
        "genero": "Mujer",
        "interaccion_id": "IT_03",
        "resumen_tematica": "Crianza con ejemplo vs gritos; crianza recibida vs crianza que quieren dar",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 1,
        "I5_normalidad": 0,
        "I6_moral": 1,
        "I7_identidad": 1,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GHV9",
        "semana": 3,
        "ciudad": "Valledupar",
        "genero": "Hombre",
        "interaccion_id": "IT_01",
        "resumen_tematica": "Comportamiento de los niños como etapas del desarrollo",
        "I1_stance": 1,
        "I2_consenso": 1,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 1,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "compleja",
    },
    {
        "grupo": "GHV9",
        "semana": 3,
        "ciudad": "Valledupar",
        "genero": "Hombre",
        "interaccion_id": "IT_02",
        "resumen_tematica": "Reflexión sobre el video: pensar antes de hablar con los hijos",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 0,
        "I4_cambio_posicion": 0,
        "I5_normalidad": 0,
        "I6_moral": 0,
        "I7_identidad": 0,
        "I8_adopcion_practica": 0,
        "nivel_interaccion": "stance-only",
    },
    {
        "grupo": "GHV9",
        "semana": 3,
        "ciudad": "Valledupar",
        "genero": "Hombre",
        "interaccion_id": "IT_03",
        "resumen_tematica": "Manejo de comportamiento difícil: hablar vs castigar físicamente",
        "I1_stance": 1,
        "I2_consenso": 0,
        "I3_desacuerdo": 1,
        "I4_cambio_posicion": 1,
        "I5_normalidad": 0,
        "I6_moral": 1,
        "I7_identidad": 1,
        "I8_adopcion_practica": 1,
        "nivel_interaccion": "compleja",
    },
]


def cargar_mensajes(df: pd.DataFrame, v_grupo: str, semana: int) -> pd.DataFrame:
    """Extrae y formatea los mensajes de un grupo-semana para el piloto."""
    sub = df[(df["v_grupo"] == v_grupo) & (df["n_week"] == semana)].copy()
    sub = sub.sort_values("datetime").reset_index(drop=True)

    sub["hora"] = pd.to_datetime(sub["datetime"]).dt.strftime("%Y-%m-%d %H:%M")

    cols = ["hora", "remitente", "id_f", "tipo", "texto"]
    return sub[cols].rename(
        columns={
            "hora": "Fecha/hora",
            "remitente": "Remitente",
            "id_f": "ID participante",
            "tipo": "Tipo",
            "texto": "Mensaje",
        }
    )


def marcar_sesiones_pp(
    msgs: pd.DataFrame,
    sesiones: pd.DataFrame,
    v_grupo: str,
    semana: int,
) -> pd.DataFrame:
    """Añade columna indicando si el mensaje cae en una sesión P-P."""
    subs = sesiones[(sesiones["v_grupo"] == v_grupo) & (sesiones["n_week"] == semana)][
        ["datetime_inicio", "datetime_fin", "sesion_id"]
    ].copy()

    if subs.empty:
        msgs["Sesion_PP"] = ""
        return msgs

    subs["datetime_inicio"] = pd.to_datetime(subs["datetime_inicio"])
    subs["datetime_fin"] = pd.to_datetime(subs["datetime_fin"])
    msgs_dt = pd.to_datetime(msgs["Fecha/hora"])

    sesion_col = []
    for dt in msgs_dt:
        matched = subs[(subs["datetime_inicio"] <= dt) & (subs["datetime_fin"] >= dt)]
        if not matched.empty:
            sesion_col.append(f"PP-{int(matched.iloc[0]['sesion_id'])}")
        else:
            sesion_col.append("")

    msgs["Sesion_PP"] = sesion_col
    return msgs


def write_sheet_mensajes(
    writer: pd.ExcelWriter,
    msgs: pd.DataFrame,
    sheet_name: str,
    info: tuple,
) -> None:
    """Escribe la hoja de mensajes con formato básico."""
    from openpyxl.styles import PatternFill

    v_grupo, semana, ciudad, genero, tema = info

    header = pd.DataFrame(
        [
            {
                "Fecha/hora": f"GRUPO: {v_grupo}",
                "Remitente": f"SEMANA: {semana}",
                "ID participante": f"CIUDAD: {ciudad}",
                "Tipo": f"GÉNERO: {genero}",
                "Mensaje": f"TEMA: {tema}",
                "Sesion_PP": "Sesion P-P (de 10a)",
            }
        ]
    )

    separador = pd.DataFrame([{col: "---" for col in msgs.columns}])
    completo = pd.concat([header, separador, msgs], ignore_index=True)
    completo.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 14

    fill_p = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    fill_pp = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
        remitente_cell = row[1]
        sesion_cell = row[5]
        if remitente_cell.value == "Participante":
            color = fill_pp if sesion_cell.value else fill_p
            for cell in row:
                cell.fill = color


def write_sheet_coding(
    writer: pd.ExcelWriter,
    its_data: list[dict],
    sheet_name: str,
) -> None:
    """Escribe una hoja de codificación con ITs pre-llenadas + filas vacías."""
    from openpyxl.styles import Font, PatternFill

    if its_data:
        pre = pd.DataFrame(its_data).reindex(columns=COLS_PLANTILLA)
    else:
        pre = pd.DataFrame(columns=COLS_PLANTILLA)

    empty = pd.DataFrame(index=range(20), columns=COLS_PLANTILLA)
    plantilla = pd.concat([pre, empty], ignore_index=True)
    plantilla.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    anchos = {
        "A": 10,
        "B": 8,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 18,
        "G": 18,
        "H": 10,
        "I": 25,
        "J": 40,
        "K": 8,
        "L": 8,
        "M": 8,
        "N": 8,
        "O": 8,
        "P": 8,
        "Q": 8,
        "R": 8,
        "S": 16,
        "T": 50,
        "U": 40,
    }
    for col_letter, width in anchos.items():
        ws.column_dimensions[col_letter].width = width

    ws["A1"].font = Font(bold=True)

    fill_compleja = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    fill_basica = PatternFill(
        start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"
    )
    fill_stance = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    nivel_fills = {
        "compleja": fill_compleja,
        "basica": fill_basica,
        "stance-only": fill_stance,
    }

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=len(pre) + 1), start=2
    ):
        nivel_cell = row[COLS_PLANTILLA.index("nivel_interaccion")]
        nivel = str(nivel_cell.value).lower() if nivel_cell.value else ""
        fill = nivel_fills.get(nivel)
        if fill:
            for cell in row:
                cell.fill = fill

    nota_row = len(plantilla) + 3
    ws.cell(row=nota_row, column=1, value=NIVEL_INSTRUCCION)
    ws.cell(row=nota_row, column=1).font = Font(italic=True, color="666666")


def write_sheet_guia(writer: pd.ExcelWriter) -> None:
    """Escribe la hoja guía de indicadores."""
    from openpyxl.styles import Alignment, Font, PatternFill

    guia_df = pd.DataFrame(GUIA_INDICADORES)[
        [
            "Nivel",
            "N",
            "Indicador",
            "Definición (Dedios et al.)",
            "Adaptación Apapachar",
            "Código en plantilla",
        ]
    ]
    guia_df.to_excel(writer, sheet_name="Guia_indicadores", index=False)

    ws = writer.sheets["Guia_indicadores"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 65
    ws.column_dimensions["F"].width = 22

    fills = {
        "Stance-only": PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        ),
        "Interacción básica": PatternFill(
            start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"
        ),
        "Interacción compleja": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
        "Específico Apapachar": PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        ),
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        nivel = row[0].value
        fill = fills.get(nivel)
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.row_dimensions[1].height = 18
    for i in range(2, 10):
        ws.row_dimensions[i].height = 80


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print("Cargando datos...")
    df = pd.read_parquet(DATA_DIR / "mensajes_preprocesados.parquet")
    sesiones = pd.read_csv(TABLES_DIR / "10a_cadenas_sesion.csv")

    out_path = TABLES_DIR / "10b_piloto_codificacion.xlsx"

    print(f"Generando {out_path}...")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        write_sheet_guia(writer)

        for info in GRUPOS_PILOTO:
            v_grupo, semana, ciudad, genero, tema = info
            sheet_name = f"{v_grupo}_s{semana}"
            print(f"  {sheet_name}...")
            msgs = cargar_mensajes(df, v_grupo, semana)
            msgs = marcar_sesiones_pp(msgs, sesiones, v_grupo, semana)
            write_sheet_mensajes(writer, msgs, sheet_name, info)

        write_sheet_coding(writer, PILOTO_ITS, "Piloto")

    print(f"\nArchivo generado: {out_path}")
    for v_grupo, semana, ciudad, genero, tema in GRUPOS_PILOTO:
        sub = df[(df["v_grupo"] == v_grupo) & (df["n_week"] == semana)]
        p = sub[sub["remitente"] == "Participante"]
        print(
            f"  {v_grupo} sem{semana} ({genero}, {ciudad}): {len(p)} msgs participantes"
        )


if __name__ == "__main__":
    main()
