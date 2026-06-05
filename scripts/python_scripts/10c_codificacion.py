"""Paso 10c: Codificación a escala — un Excel por chunk
=======================================================
Genera outputs/tables/10c_chunkN.xlsx (N=1..6) para cada uno de los
6 chunks de 8 grupos representativos (semanas 1-2, 3-4, ..., 11-12).

Cada Excel tiene la misma estructura que el piloto (10b):
  - Guia_indicadores
  - Una hoja de mensajes por grupo (nombre: GRUPO_sN)
  - Una hoja de codificación (nombre: Chunk_N)

Los grupos representativos se seleccionaron en el Paso 10b:
  - Semanas impares -> grupo de mujeres más activo por ciudad
  - Semanas pares  -> grupo de hombres más activo por ciudad
  Ver outputs/tables/10b_grupos_representativos.csv

Input:
  - data/clean/mensajes_preprocesados.parquet
  - outputs/tables/10a_cadenas_sesion.csv

Output:
  - outputs/tables/10c_chunk1.xlsx … outputs/tables/10c_chunk6.xlsx
"""

from __future__ import annotations

import pandas as pd
from config_loader import PROJECT_ROOT, TABLES_DIR, cfg

# ---------------------------------------------------------------------------
# Configuración (valores leídos de config.yaml)
# ---------------------------------------------------------------------------
PROJECT_NAME = cfg["project"]["name"]
ADAPTATION_KEY = f"Adaptación {PROJECT_NAME}"

COL_GROUP = cfg["data"]["columns"]["group_id"]
COL_WEEK = cfg["data"]["columns"]["week_number"]
COL_DATETIME = cfg["data"]["columns"]["datetime"]
COL_SENDER = cfg["data"]["columns"]["sender"]
COL_PARTICIPANT_ID = cfg["data"]["columns"]["participant_id"]
COL_TYPE = cfg["data"]["columns"]["message_type"]
COL_TEXT = cfg["data"]["columns"]["message_text"]
COL_THEME = cfg["data"]["columns"]["theme"]

# ---------------------------------------------------------------------------
# Guía de indicadores (construida desde config.yaml)
# ---------------------------------------------------------------------------
GUIA_INDICADORES = [
    {
        "Nivel": ind["level"],
        "N": ind["id"],
        "Indicador": ind["name"],
        "Definición (Dedios et al.)": ind["definition_dedios"].strip(),
        ADAPTATION_KEY: ind["adaptation"].strip(),
        "Código en plantilla": ind["template_code"],
    }
    for ind in cfg["coding"]["indicators"]
]

_indicator_codes = [ind["template_code"] for ind in cfg["coding"]["indicators"]]

COLS_PLANTILLA = [
    "grupo",
    "semana",
    "ciudad",
    "genero",
    "interaccion_id",
    "datetime_inicio",
    "datetime_fin",
    "n_mensajes",
    "ids_participantes",
    "resumen_tematica",
    *_indicator_codes,
    "nivel_interaccion",
    "citas_relevantes",
    "notas",
]

NIVEL_INSTRUCCION = (
    "nivel_interaccion: 'stance-only' si solo I1=1; "
    "'basica' si I2/I3/I4 ≥ 1; 'compleja' si I5/I6/I7 ≥ 1. Tomar el más alto."
)

# ---------------------------------------------------------------------------
# Chunks: grupos representativos e ITs codificadas
# ---------------------------------------------------------------------------
# Formato grupos: (v_grupo, semana, ciudad, genero)
# Formato ITs: lista de dicts con claves de COLS_PLANTILLA

CHUNKS: list[dict] = [
    # -----------------------------------------------------------------------
    # [ADAPTAR] Añadir un dict por cada periodo de análisis.
    #
    # Cada chunk representa un bloque de semanas / periodo a codificar.
    # Los grupos representativos se seleccionan con el script 10b.
    # Las ITs (interacciones P-P) se identifican y codifican con Claude Code.
    #
    # Estructura de cada entrada:
    #   num      : número del chunk (1, 2, 3 …)
    #   semanas  : etiqueta del periodo (ej. "1-2", "3-4")
    #   grupos   : lista de tuplas (v_grupo, semana, ciudad, genero)
    #   its      : lista de ITs codificadas; cada IT debe incluir las
    #              claves definidas en COLS_PLANTILLA arriba.
    #
    # Ejemplo mínimo (descomentar y adaptar):
    # {
    #     "num": 1,
    #     "semanas": "1-2",
    #     "grupos": [
    #         ("GRUPO_A", 1, "Ciudad1", "Mujeres"),
    #         ("GRUPO_B", 2, "Ciudad1", "Hombres"),
    #     ],
    #     "its": [
    #         {
    #             "grupo": "GRUPO_A",
    #             "semana": 1,
    #             "ciudad": "Ciudad1",
    #             "genero": "Mujer",
    #             "interaccion_id": "IT_01",
    #             "datetime_inicio": "2025-01-01 10:00",
    #             "datetime_fin": "2025-01-01 11:00",
    #             "n_mensajes": 5,
    #             "ids_participantes": "P001|P002|P003",
    #             "resumen_tematica": "Tema de la interacción",
    #             # Indicadores DEDIOS (0 o 1):
    #             "I1_stance": 1,
    #             "I2_consenso": 0,
    #             # … resto de indicadores del config …
    #             "nivel_interaccion": "stance-only",
    #             "citas_relevantes": "",
    #             "notas": "",
    #         },
    #     ],
    # },
]

# ---------------------------------------------------------------------------
# Funciones de escritura (idénticas a 10b)
# ---------------------------------------------------------------------------


def get_tema(df: pd.DataFrame, v_grupo: str, semana: int) -> str:
    """Obtiene el tema de una sesión desde el DataFrame de mensajes."""
    sub = df[(df[COL_GROUP] == v_grupo) & (df[COL_WEEK] == semana)]
    if COL_THEME in sub.columns and not sub.empty:
        vals = sub[COL_THEME].dropna()
        return vals.iloc[0] if not vals.empty else ""
    return ""


def cargar_mensajes(df: pd.DataFrame, v_grupo: str, semana: int) -> pd.DataFrame:
    """Extrae y formatea los mensajes de un grupo-semana."""
    sub = df[(df[COL_GROUP] == v_grupo) & (df[COL_WEEK] == semana)].copy()
    sub = sub.sort_values(COL_DATETIME).reset_index(drop=True)
    sub["hora"] = pd.to_datetime(sub[COL_DATETIME]).dt.strftime("%Y-%m-%d %H:%M")
    cols = ["hora", COL_SENDER, COL_PARTICIPANT_ID, COL_TYPE, COL_TEXT]
    return sub[cols].rename(
        columns={
            "hora": "Fecha/hora",
            COL_SENDER: "Remitente",
            COL_PARTICIPANT_ID: "ID participante",
            COL_TYPE: "Tipo",
            COL_TEXT: "Mensaje",
        }
    )


def marcar_sesiones_pp(
    msgs: pd.DataFrame,
    sesiones: pd.DataFrame,
    v_grupo: str,
    semana: int,
) -> pd.DataFrame:
    """Añade columna indicando si el mensaje cae en una sesión P-P."""
    subs = sesiones[(sesiones[COL_GROUP] == v_grupo) & (sesiones[COL_WEEK] == semana)][
        ["datetime_inicio", "datetime_fin", "sesion_id"]
    ].copy()

    if subs.empty:
        msgs["Sesion_PP"] = ""
        return msgs

    subs["datetime_inicio"] = pd.to_datetime(subs["datetime_inicio"])
    subs["datetime_fin"] = pd.to_datetime(subs["datetime_fin"])
    msgs_dt = pd.to_datetime(msgs["Fecha/hora"])

    sesion_col = []
    for dt in msgs_dt:
        matched = subs[(subs["datetime_inicio"] <= dt) & (subs["datetime_fin"] >= dt)]
        sesion_col.append(
            f"PP-{int(matched.iloc[0]['sesion_id'])}" if not matched.empty else ""
        )

    msgs["Sesion_PP"] = sesion_col
    return msgs


def write_sheet_mensajes(
    writer: pd.ExcelWriter,
    msgs: pd.DataFrame,
    sheet_name: str,
    info: tuple,
) -> None:
    """Escribe la hoja de mensajes con formato y color-coding P-P."""
    from openpyxl.styles import PatternFill

    v_grupo, semana, ciudad, genero, tema = info

    header = pd.DataFrame(
        [
            {
                "Fecha/hora": f"GRUPO: {v_grupo}",
                "Remitente": f"SEMANA: {semana}",
                "ID participante": f"CIUDAD: {ciudad}",
                "Tipo": f"GÉNERO: {genero}",
                "Mensaje": f"TEMA: {tema}",
                "Sesion_PP": "Sesion P-P (de 10a)",
            }
        ]
    )
    separador = pd.DataFrame([{col: "---" for col in msgs.columns}])
    completo = pd.concat([header, separador, msgs], ignore_index=True)
    completo.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 14

    fill_p = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    fill_pp = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
        if row[1].value == "Participante":
            color = fill_pp if row[5].value else fill_p
            for cell in row:
                cell.fill = color


def write_sheet_coding(
    writer: pd.ExcelWriter,
    its_data: list[dict],
    sheet_name: str,
) -> None:
    """Escribe la hoja de codificación con ITs pre-llenadas + filas vacías."""
    from openpyxl.styles import Font, PatternFill

    pre = (
        pd.DataFrame(its_data).reindex(columns=COLS_PLANTILLA)
        if its_data
        else pd.DataFrame(columns=COLS_PLANTILLA)
    )
    empty = pd.DataFrame(index=range(20), columns=COLS_PLANTILLA)
    plantilla = pd.concat([pre, empty], ignore_index=True)
    plantilla.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    anchos = {
        "A": 10,
        "B": 8,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 18,
        "G": 18,
        "H": 10,
        "I": 25,
        "J": 40,
        "K": 8,
        "L": 8,
        "M": 8,
        "N": 8,
        "O": 8,
        "P": 8,
        "Q": 8,
        "R": 8,
        "S": 16,
        "T": 50,
        "U": 40,
    }
    for col_letter, width in anchos.items():
        ws.column_dimensions[col_letter].width = width

    ws["A1"].font = Font(bold=True)

    nivel_fills = {
        "compleja": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
        "basica": PatternFill(
            start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"
        ),
        "stance-only": PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        ),
    }
    idx_nivel = COLS_PLANTILLA.index("nivel_interaccion")
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=len(pre) + 1), start=2
    ):
        nivel = str(row[idx_nivel].value).lower() if row[idx_nivel].value else ""
        fill = nivel_fills.get(nivel)
        if fill:
            for cell in row:
                cell.fill = fill

    nota_row = len(plantilla) + 3
    ws.cell(row=nota_row, column=1, value=NIVEL_INSTRUCCION)
    ws.cell(row=nota_row, column=1).font = Font(italic=True, color="666666")


def write_sheet_guia(writer: pd.ExcelWriter) -> None:
    """Escribe la hoja guía de indicadores."""
    from openpyxl.styles import Alignment, Font, PatternFill

    guia_df = pd.DataFrame(GUIA_INDICADORES)[
        [
            "Nivel",
            "N",
            "Indicador",
            "Definición (Dedios et al.)",
            ADAPTATION_KEY,
            "Código en plantilla",
        ]
    ]
    guia_df.to_excel(writer, sheet_name="Guia_indicadores", index=False)

    ws = writer.sheets["Guia_indicadores"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 65
    ws.column_dimensions["F"].width = 22

    _extra_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    fills = {
        "Stance-only": PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        ),
        "Interacción básica": PatternFill(
            start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"
        ),
        "Interacción compleja": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
    }
    for _ind in cfg["coding"]["indicators"]:
        if _ind["level"] not in fills:
            fills[_ind["level"]] = _extra_fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        nivel = row[0].value
        fill = fills.get(nivel)
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.row_dimensions[1].height = 18
    for i in range(2, 10):
        ws.row_dimensions[i].height = 80


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print("Cargando datos...")
    df = pd.read_parquet(
        PROJECT_ROOT / cfg["data"]["intermediate"]["preprocessed_messages"]
    )
    sesiones = pd.read_csv(TABLES_DIR / "10a_cadenas_sesion.csv")

    for chunk in CHUNKS:
        n = chunk["num"]
        out_path = TABLES_DIR / f"10c_chunk{n}.xlsx"
        print(
            f"\nGenerando chunk {n} (semanas {chunk['semanas']}) -> {out_path.name}..."
        )

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            write_sheet_guia(writer)

            for v_grupo, semana, ciudad, genero in chunk["grupos"]:
                tema = get_tema(df, v_grupo, semana)
                info = (v_grupo, semana, ciudad, genero, tema)
                sheet_name = f"{v_grupo}_s{semana}"
                print(f"  {sheet_name}...")
                msgs = cargar_mensajes(df, v_grupo, semana)
                msgs = marcar_sesiones_pp(msgs, sesiones, v_grupo, semana)
                write_sheet_mensajes(writer, msgs, sheet_name, info)

            write_sheet_coding(writer, chunk["its"], f"Chunk_{n}")

        n_its = len(chunk["its"])
        print(f"  -> {n_its} ITs codificadas")

    print("\nListo.")


if __name__ == "__main__":
    main()
